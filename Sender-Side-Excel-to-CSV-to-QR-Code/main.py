import openpyxl
from openpyxl import load_workbook
import csv
from csv import reader
import qrcode
import os
import os.path


files_to_read = [("126sqnexcel","FileName"),( "flyingprogramme","Filename2"),("sqnexcel","FileName3"),( "sarprogramme","Filename4")]

for (file, _) in files_to_read:
  if os.path.exists(file + '.csv'):
    os.remove(file + '.csv')
  else:
    pass

  workbook = load_workbook(filename = 'sqnexcel.xlsx', data_only=True) #data_only=True prevent formula printed
  sheet_obj = workbook[file]
  m_row = sheet_obj.max_row
 
# Loop will print all values
# of first column
            
  for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    data = cell_obj.value
    #data is what's needed in each file
    #print(data)
          

    with open(file + '.csv', 'a') as csvfile1:
    #open '.csv' and add
      fieldNames = ['default']
      writer = csv.DictWriter(csvfile1, fieldnames=fieldNames)
      writer.writerow({'default': data})
      csvfile1.close()
  #Section above export excel sheet as csv

  with open(file + '.csv') as csvfile:
    reader= csv.reader(csvfile)

    qr = qrcode.QRCode(
      version=1,
      error_correction=qrcode.constants.ERROR_CORRECT_L,
      box_size=3,
      border=4,
    )


    for i, row in enumerate(reader):
      labeldata = row[0] #note bug: excel file cannot have rows with no values else error. replaced with -------
      print (labeldata)

      qr.add_data(labeldata + "\n") #here, add is wrong
      qr.make(fit=True)

      img = qr.make_image()
      img.save(file + ".jpg")
      #img.save("sqnexcel.jpg".format(i))
    
      #wb = load_workbook(filename = 'sqnexcel.xlsx')
      #ws = wb[file]
      sheet_obj._images = [] #clear all images in sheet
      img = openpyxl.drawing.image.Image(file + '.jpg')
      sheet_obj.add_image(img,'E2')
      workbook.save('sqnexcel.xlsx')
      workbook.close()
    

os.remove('flyingprogramme.csv')
os.remove('sarprogramme.csv')
os.remove('sqnexcel.csv')
os.remove('126sqnexcel.csv')
os.remove('flyingprogramme.jpg')
os.remove('sarprogramme.jpg')
os.remove('126sqnexcel.jpg')
os.remove('sqnexcel.jpg')
#clean folder of unnecessary files
